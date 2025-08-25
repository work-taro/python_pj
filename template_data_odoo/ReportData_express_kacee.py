import pymysql
import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook


def clean_column(val):
    # แก้ค่าเป็น 0.00 → '-'
    try:
        return '-' if float(val) == 0.00 else val
    except:
        return val


def get_customer_code():
    connection = pymysql.connect(
        host='localhost',
        user='root',
        password='',
        db='odoo_kacee',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    query = "SELECT CUSCOD, NEWCUSCOD FROM customer_kacee WHERE CUSCOD IS NOT NULL AND NEWCUSCOD IS NOT NULL"

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()
        df_map = pd.DataFrame(result)
    connection.close()
    return df_map


def get_vendor_code():
    connection = pymysql.connect(
        host='localhost',
        user='root',
        password='',
        db='odoo_kacee',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    query = "SELECT SUPCOD, new_code FROM vendor_kacee WHERE SUPCOD IS NOT NULL AND new_code IS NOT NULL"

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()
        df_map = pd.DataFrame(result)
    connection.close()
    return df_map


def ProcessFile_1A2(df):
    def create_column_n(row):
        if row['K'] != '-':
            return f"Invoice = {row['J']} รับชำระแล้ว = {row['K']}"
        else:
            return ''

    def create_column_m(row):
        return ''

    # เลือกเฉพาะคอลัมน์ 0-11 (เท่าที่มีจริง)
    df = df[df.columns[:12]]

    # แสดงเฉพาะคอลัมน์ D–L (index 3 ถึง 11)
    df = df.iloc[:, 3:12]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

    # ────── เงื่อนไขการกรอง ──────
    df = df[~df['L'].astype(str).str.strip().eq('---------------')]
    df = df[~df['K'].astype(str).str.strip().eq('ใบ')]
    df = df[~df['I'].astype(str).str.startswith('R')]

    # ────── เติมข้อมูลคอลัมน์ D, E ที่ว่าง ด้วยค่าเดิมจากแถวก่อนหน้า ──────
    df[['D', 'E']] = df[['D', 'E']].ffill()

    df['K'] = df['K'].apply(clean_column)
    df['M'] = df.apply(create_column_m, axis=1)
    df['N'] = df.apply(create_column_n, axis=1)

    # ────── ลบแถวที่ F ถึง L ว่างทั้งหมด ──────
    df = df.dropna(subset=['F', 'H', 'I', 'J', 'K', 'L'], how='all')
    df = df.drop(columns=['G'])

    # ลบ row ที่มี cell ใดก็ตามขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก (header เก่า)
    df = df.iloc[2:].reset_index(drop=True)

    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'Partner', 'Partner Code', 'Invoice Date', 'Payment Reference',
        'Sale Person', 'text_1', 'text_2', 'Total', '', 'Note'
    ]

    # ────── ดึงข้อมูล Mapping จาก Database ──────
    customer_map_df = get_customer_code()
    customer_map = dict(zip(customer_map_df['CUSCOD'], customer_map_df['NEWCUSCOD']))

    # ────── Map ค่า Partner Code ──────
    def map_customer(code):
        return customer_map.get(code, code)

    df['Original Code'] = df['Partner Code']  # เก็บค่าเดิมไว้ก่อน
    df['Partner Code'] = df['Partner Code'].apply(map_customer)  # แมปรหัสใหม่
    df['IsMissing'] = df['Partner Code'] == df['Original Code']  # ถ้ายังเหมือนเดิม = ไม่เจอ
    df = df.drop(columns=['Original Code'])  # ไม่ต้องใช้แล้ว

    # ────── ลบแถวที่ว่างจริง ๆ ──────
    df = df.dropna(subset=['Invoice Date', 'Payment Reference', 'Total'], how='all')

    # ────── Export to Excel ──────
    export_path = 'C:/Users/kc/Desktop/data_csv_to_template/exports/1A2.xlsx'
    df.drop(columns=['IsMissing']).to_excel(export_path, index=False, engine='openpyxl')

    # ────── ใส่สีแดงให้ Partner Code ที่หาไม่เจอ ──────
    wb = load_workbook(export_path)
    ws = wb.active

    # หา index ของคอลัมน์ 'Partner Code'
    partner_code_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Partner Code':
            partner_code_col_idx = idx
            break

    # ใส่สีแดงแถวที่ไม่เจอ
    for i, is_missing in enumerate(df['IsMissing'], start=2):  # เริ่ม row 2
        if is_missing:
            cell = ws.cell(row=i, column=partner_code_col_idx)
            cell.font = Font(color="FF0000")  # สีแดง

    wb.save(export_path)
    print("✅ Export Files 1A2 : เรียบร้อย")


def ProcessFile_116(df):
    # สร้างคอลัมน์ M ตามเงื่อนไข
    def create_column_n(row):
        if row['J'] != '-':
            return f"AI = {row['J']} ,ตัดชำระ = {row['J']}"
        else:
            return ''

    def create_column_m(row):
        return ''

    # เลือกเฉพาะคอลัมน์ 0-11 (เท่าที่มีจริง)
    df = df[df.columns[:12]]

    # แสดงเฉพาะคอลัมน์ C–L (index 2 ถึง 11)
    df = df.iloc[:, 2:12]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

    # ────── เงื่อนไขการกรอง ──────

    df = df[~df['L'].astype(str).str.strip().eq('---------------')]
    df = df[~df['L'].astype(str).str.strip().eq('===============')]

    df = df[~df['J'].astype(str).str.strip().eq('ใบ')]
    df = df[~df['H'].astype(str).str.startswith('R')]

    # ────── เติมข้อมูลคอลัมน์ D, E ที่ว่าง ด้วยค่าเดิมจากแถวก่อนหน้า ──────
    df[['C', 'D']] = df[['C', 'D']].ffill()

    df['J'] = df['J'].apply(clean_column)
    df['M'] = df.apply(create_column_m, axis=1)
    df['N'] = df.apply(create_column_n, axis=1)

    # ────── ลบแถวที่ F ถึง L ว่างทั้งหมด (แถวหัวหรือช่องว่าง) ──────
    df = df.dropna(subset=['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], how='all')
    df = df.drop(columns=['F'])

    # ลบ row ที่มี cell ใดก็ตามในแถวขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก
    df = df.iloc[2:].reset_index(drop=True)

    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'Partner', 'Partner Code', 'Invoice Date', 'Payment Reference',
        'Sale Person', 'text_1', 'text_2', 'Total', 'Total VAT', '', 'Note'
    ]
    df = df[~df['Sale Person'].astype(str).str.startswith('VT')]
    df = df[~df['Sale Person'].astype(str).str.startswith('IW')]

    # ────── ดึงข้อมูล Mapping จาก Database ──────
    customer_map_df = get_customer_code()
    customer_map = dict(zip(customer_map_df['CUSCOD'], customer_map_df['NEWCUSCOD']))

    # ────── Map ค่า Partner Code ──────
    def map_customer(code):
        return customer_map.get(code, code)

    df['Original Code'] = df['Partner Code']  # เก็บค่าเดิมไว้ก่อน
    df['Partner Code'] = df['Partner Code'].apply(map_customer)  # แมปรหัสใหม่
    df['IsMissing'] = df['Partner Code'] == df['Original Code']  # ถ้ายังเหมือนเดิม = ไม่เจอ
    df = df.drop(columns=['Original Code'])  # ไม่ต้องใช้แล้ว

    # ────── ลบแถวที่ว่างจริง ๆ ──────
    df = df.dropna(subset=['Invoice Date', 'Payment Reference', 'Total'], how='all')

    # ────── Export to Excel ──────
    export_path = 'C:/Users/kc/Desktop/data_csv_to_template/exports/116.xlsx'
    df.drop(columns=['IsMissing']).to_excel(export_path, index=False, engine='openpyxl')

    # ────── ใส่สีแดงให้ Partner Code ที่หาไม่เจอ ──────
    wb = load_workbook(export_path)
    ws = wb.active

    # หา index ของคอลัมน์ 'Partner Code'
    partner_code_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Partner Code':
            partner_code_col_idx = idx
            break

    # ใส่สีแดงแถวที่ไม่เจอ
    for i, is_missing in enumerate(df['IsMissing'], start=2):  # เริ่ม row 2
        if is_missing:
            cell = ws.cell(row=i, column=partner_code_col_idx)
            cell.font = Font(color="FF0000")  # สีแดง

    wb.save(export_path)
    print("✅ Export Files 116 : เรียบร้อย")


def ProcessFile_3C(df):
    def create_column_r(row):
        if pd.notna(row['P']):
            return f"{row['O']}_{row['P']}"
        else:
            return f"{row['O']}"

    def create_column_q(row):
        return 'Recive Money'

    # เลือกเฉพาะคอลัมน์ A–P
    df = df[df.columns[:16]]

    # แสดงเฉพาะคอลัมน์ C–P
    df = df.iloc[:, 2:16]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

    # ────── เงื่อนไขการกรอง ──────
    df = df[~df['L'].astype(str).str.strip().eq('---------------')]
    df = df[~df['K'].astype(str).str.strip().eq('ใบ')]

    # ────── เติมข้อมูลคอลัมน์ D, E ที่ว่าง ด้วยค่าเดิมจากแถวก่อนหน้า ──────
    df[['C', 'D']] = df[['C', 'D']].ffill()

    # ────── ลบแถวที่ E ถึง O ว่างทั้งหมด (แถวหัวหรือช่องว่าง) ──────
    df = df.dropna(subset=['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O'], how='all')
    df['Q'] = df.apply(create_column_q, axis=1)
    df['R'] = df.apply(create_column_r, axis=1)

    df = df.drop(columns=['K'])
    df = df.drop(columns=['M'])
    df = df.drop(columns=['O'])
    df = df.drop(columns=['P'])

    # ลบ row ที่มี cell ใดก็ตามในแถวขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก
    df = df.iloc[1:].reset_index(drop=True)

    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'text_1', 'text_2', 'PDC Date', 'Due Date',
        'Cheque Reference', 'Bank', 'Bank Branch', 'Partner Code', 'Amount', 'Journal', 'Payment Type', 'Memo'
    ]
    # ────── ดึงข้อมูล Mapping จาก Database ──────
    customer_map_df = get_customer_code()
    customer_map = dict(zip(customer_map_df['CUSCOD'], customer_map_df['NEWCUSCOD']))

    # ────── Map ค่า Partner Code ──────
    def map_customer(code):
        return customer_map.get(code, code)

    df['Original Code'] = df['Partner Code']  # เก็บค่าเดิมไว้ก่อน
    df['Partner Code'] = df['Partner Code'].apply(map_customer)  # แมปรหัสใหม่
    df['IsMissing'] = df['Partner Code'] == df['Original Code']  # ถ้ายังเหมือนเดิม = ไม่เจอ
    df = df.drop(columns=['Original Code'])  # ไม่ต้องใช้แล้ว

    # ────── Export to Excel ──────
    export_path = 'C:/Users/kc/Desktop/data_csv_to_template/exports/3C.xlsx'
    df.drop(columns=['IsMissing']).to_excel(export_path, index=False, engine='openpyxl')

    # ────── ใส่สีแดงให้ Partner Code ที่หาไม่เจอ ──────
    wb = load_workbook(export_path)
    ws = wb.active

    # หา index ของคอลัมน์ 'Partner Code'
    partner_code_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Partner Code':
            partner_code_col_idx = idx
            break

    # ใส่สีแดงแถวที่ไม่เจอ
    for i, is_missing in enumerate(df['IsMissing'], start=2):  # เริ่ม row 2
        if is_missing:
            cell = ws.cell(row=i, column=partner_code_col_idx)
            cell.font = Font(color="FF0000")  # สีแดง

    wb.save(export_path)
    print("✅ Export Files 3C : เรียบร้อย")


def ProcessFile_2A2(df):
    # สร้างคอลัมน์ M ตามเงื่อนไข
    def create_column_n(row):
        if row['L'] != '-':
            return f"Bill = {row['K']} ,จ่ายชำระแล้ว = {row['L']}"
        else:
            return ''

    # เลือกเฉพาะคอลัมน์ 0-11 (เท่าที่มีจริง)
    df = df[df.columns[:14]]

    # แสดงเฉพาะคอลัมน์ D–L (index 3 ถึง 11)
    df = df.iloc[:, 3:14]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']

    # ────── เงื่อนไขการกรอง ──────

    df = df[~df['M'].astype(str).str.strip().eq('---------------')]
    df = df[~df['G'].astype(str).str.strip().eq('ใบ')]
    df = df[~df['H'].astype(str).str.strip().eq('ใบ')]

    # ────── เติมข้อมูลคอลัมน์ D, E ที่ว่าง ด้วยค่าเดิมจากแถวก่อนหน้า ──────
    df[['D', 'E']] = df[['D', 'E']].ffill()

    df['L'] = df['L'].apply(clean_column)
    df['N'] = df.apply(create_column_n, axis=1)

    # # ────── ลบแถวที่ F ถึง L ว่างทั้งหมด (แถวหัวหรือช่องว่าง) ──────
    df = df.dropna(subset=['G', 'H', 'I', 'J', 'K', 'L', 'M'], how='all')
    df = df.drop(columns=['F'])
    df = df.drop(columns=['H'])

    # ลบ row ที่มี cell ใดก็ตามในแถวขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก
    df = df.iloc[2:].reset_index(drop=True)

    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'Partner', 'Partner Code', 'Bill Date', 'Bill No',
        'text_1', 'text_2', 'text_3', 'Total', 'Note'
    ]

    df = df[~df['text_1'].astype(str).str.startswith('PS')]
    df = df[~df['text_1'].astype(str).str.startswith('PN')]
    # ────── ดึงข้อมูล Mapping จาก Database ──────
    vendor_map_df = get_vendor_code()
    vendor_map = dict(zip(vendor_map_df['SUPCOD'], vendor_map_df['new_code']))

    # ────── Map ค่า Partner Code ──────
    def map_vendor(code):
        return vendor_map.get(code, code)

    df['Original Code'] = df['Partner Code']  # เก็บค่าเดิมไว้ก่อน
    df['Partner Code'] = df['Partner Code'].apply(map_vendor)  # แมปรหัสใหม่
    df['IsMissing'] = df['Partner Code'] == df['Original Code']  # ถ้ายังเหมือนเดิม = ไม่เจอ
    df = df.drop(columns=['Original Code'])  # ไม่ต้องใช้แล้ว

    # ────── Export to Excel ──────
    export_path = 'C:/Users/kc/Desktop/data_csv_to_template/exports/2A2.xlsx'
    df.drop(columns=['IsMissing']).to_excel(export_path, index=False, engine='openpyxl')

    # ────── ใส่สีแดงให้ Partner Code ที่หาไม่เจอ ──────
    wb = load_workbook(export_path)
    ws = wb.active

    # หา index ของคอลัมน์ 'Partner Code'
    partner_code_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Partner Code':
            partner_code_col_idx = idx
            break

    # ใส่สีแดงแถวที่ไม่เจอ
    for i, is_missing in enumerate(df['IsMissing'], start=2):  # เริ่ม row 2
        if is_missing:
            cell = ws.cell(row=i, column=partner_code_col_idx)
            cell.font = Font(color="FF0000")  # สีแดง

    wb.save(export_path)
    print("✅ Export Files 2A2 : เรียบร้อย")


def ProcessFile_215(df):
    # สร้างคอลัมน์ M ตามเงื่อนไข
    def create_column_s(row):
        if row['O'] != '-':
            return f"จ่ายมัดจำ = {row['N']} , ตัดชำระ = {row['O']}"
        else:
            return ''

    def create_column_r(row):
        return ''

    # เลือกเฉพาะคอลัมน์ A-Q (เท่าที่มีจริง)
    df = df[df.columns[:17]]

    # แสดงเฉพาะคอลัมน์ D–Q (index 3 ถึง 17)
    df = df.iloc[:, 3:17]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']

    # ────── เงื่อนไขการกรอง ──────
    df = df[~df['P'].astype(str).str.strip().eq('---------------')]
    df = df[~df['G'].astype(str).str.strip().eq('ใบ')]
    df = df[~df['I'].astype(str).str.strip().eq('ใบ')]

    # ────── เติมข้อมูลคอลัมน์ D, E ที่ว่าง ด้วยค่าเดิมจากแถวก่อนหน้า ──────
    df[['D', 'E']] = df[['D', 'E']].ffill()

    df['O'] = df['O'].apply(clean_column)
    df['R'] = df.apply(create_column_r, axis=1)
    df['S'] = df.apply(create_column_s, axis=1)

    # ────── ลบแถวที่ F ถึง L ว่างทั้งหมด (แถวหัวหรือช่องว่าง) ──────
    df = df.dropna(subset=['J', 'L', 'P', 'Q'], how='all')
    df = df.drop(columns=['F'])
    df = df.drop(columns=['G'])
    df = df.drop(columns=['H'])
    df = df.drop(columns=['I'])
    df = df.drop(columns=['K'])

    # ลบ row ที่มี cell ใดก็ตามในแถวขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก
    df = df.iloc[1:].reset_index(drop=True)

    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'Partner', 'Partner Code', 'Invoice Date', 'Payment Reference',
        'text_1', 'text_2', 'text_3', 'Total', 'text_4', '', 'Note'
    ]
    # ────── ดึงข้อมูล Mapping จาก Database ──────
    vendor_map_df = get_vendor_code()
    vendor_map = dict(zip(vendor_map_df['SUPCOD'], vendor_map_df['new_code']))

    # ────── Map ค่า Partner Code ──────
    def map_vendor(code):
        return vendor_map.get(code, code)

    df['Original Code'] = df['Partner Code']  # เก็บค่าเดิมไว้ก่อน
    df['Partner Code'] = df['Partner Code'].apply(map_vendor)  # แมปรหัสใหม่
    df['IsMissing'] = df['Partner Code'] == df['Original Code']  # ถ้ายังเหมือนเดิม = ไม่เจอ
    df = df.drop(columns=['Original Code'])  # ไม่ต้องใช้แล้ว

    # ────── Export to Excel ──────
    export_path = 'C:/Users/kc/Desktop/data_csv_to_template/exports/215.xlsx'
    df.drop(columns=['IsMissing']).to_excel(export_path, index=False, engine='openpyxl')

    # ────── ใส่สีแดงให้ Partner Code ที่หาไม่เจอ ──────
    wb = load_workbook(export_path)
    ws = wb.active

    # หา index ของคอลัมน์ 'Partner Code'
    partner_code_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Partner Code':
            partner_code_col_idx = idx
            break

    # ใส่สีแดงแถวที่ไม่เจอ
    for i, is_missing in enumerate(df['IsMissing'], start=2):  # เริ่ม row 2
        if is_missing:
            cell = ws.cell(row=i, column=partner_code_col_idx)
            cell.font = Font(color="FF0000")  # สีแดง

    wb.save(export_path)

    print("✅ Export Files 215 : เรียบร้อย")


def ProcessFile_3D(df):
    # สร้างคอลัมน์ O ตามเงื่อนไข
    def create_column_o(row):
        return 'Send Money'

    # เลือกเฉพาะคอลัมน์ A-N (เท่าที่มีจริง)
    df = df[df.columns[:14]]

    # แสดงเฉพาะคอลัมน์ C–N (index 2 ถึง 14)
    df = df.iloc[:, 2:14]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

    # ────── เงื่อนไขการกรอง ──────
    df = df[~df['H'].astype(str).str.strip().eq('-------------')]
    df = df[~df['N'].astype(str).str.strip().eq('ใบ')]

    # # ────── เติมข้อมูลคอลัมน์ D, E ที่ว่าง ด้วยค่าเดิมจากแถวก่อนหน้า ──────
    df[['C', 'D']] = df[['C', 'D']].ffill()

    df['N'] = df['N'].apply(clean_column)
    df['O'] = df.apply(create_column_o, axis=1)
    #
    # ────── ลบแถวที่ F ถึง L ว่างทั้งหมด (แถวหัวหรือช่องว่าง) ──────
    df = df.dropna(subset=['E', 'F', 'G', 'J', 'K'], how='all')
    df = df.drop(columns=['H'])
    df = df.drop(columns=['I'])
    df = df.drop(columns=['L'])
    df = df.drop(columns=['M'])

    # ลบ row ที่มี cell ใดก็ตามในแถวขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก
    df = df.iloc[1:].reset_index(drop=True)
    #
    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'text_1', 'text_2', 'PDC Date', 'Due Date',
        'Cheque Reference', 'Partner Code', 'Memo', 'PDC Amount', 'Payment Type'
    ]
    # ────── ดึงข้อมูล Mapping จาก Database ──────
    vendor_map_df = get_vendor_code()
    vendor_map = dict(zip(vendor_map_df['SUPCOD'], vendor_map_df['new_code']))

    # ────── Map ค่า Partner Code ──────
    def map_vendor(code):
        return vendor_map.get(code, code)

    df['Original Code'] = df['Partner Code']  # เก็บค่าเดิมไว้ก่อน
    df['Partner Code'] = df['Partner Code'].apply(map_vendor)  # แมปรหัสใหม่
    df['IsMissing'] = df['Partner Code'] == df['Original Code']  # ถ้ายังเหมือนเดิม = ไม่เจอ
    df = df.drop(columns=['Original Code'])  # ไม่ต้องใช้แล้ว

    # ────── Export to Excel ──────
    export_path = 'C:/Users/kc/Desktop/data_csv_to_template/exports/3D.xlsx'
    df.drop(columns=['IsMissing']).to_excel(export_path, index=False, engine='openpyxl')

    # ────── ใส่สีแดงให้ Partner Code ที่หาไม่เจอ ──────
    wb = load_workbook(export_path)
    ws = wb.active

    # หา index ของคอลัมน์ 'Partner Code'
    partner_code_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Partner Code':
            partner_code_col_idx = idx
            break

    # ใส่สีแดงแถวที่ไม่เจอ
    for i, is_missing in enumerate(df['IsMissing'], start=2):  # เริ่ม row 2
        if is_missing:
            cell = ws.cell(row=i, column=partner_code_col_idx)
            cell.font = Font(color="FF0000")  # สีแดง

    wb.save(export_path)
    print("✅ Export Files 3D : เรียบร้อย")


def ProcessFile_TB(df):
    # ไฟล์ 566(ต้นฉบับ)
    # เลือกเฉพาะคอลัมน์ A-G(เท่าที่มีจริง)
    df = df[df.columns[:6]]

    # แสดงเฉพาะคอลัมน์ B–F (index 1 ถึง 5)
    df = df.iloc[:, 1:6]

    # ลบแถวที่ว่างทั้งหมดในช่วงคอลัมน์นี้
    df = df.dropna(how='all')

    # ตั้งชื่อคอลัมน์
    df.columns = ['B', 'C', 'D', 'E', 'F']

    # ────── เงื่อนไขการกรอง ──────
    df = df[~df['E'].astype(str).str.strip().eq('-----------------')]
    df = df[~df['D'].astype(str).str.strip().eq('บัญชี')]

    df['E'] = df['E'].apply(clean_column)
    df['F'] = df['F'].apply(clean_column)

    # # ────── ลบแถวที่ F ถึง L ว่างทั้งหมด (แถวหัวหรือช่องว่าง) ──────
    df = df.dropna(subset=['C', 'D', 'E', 'F'], how='all')
    df = df.drop(columns=['D'])

    # ลบ row ที่มี cell ใดก็ตามในแถวขึ้นต้นด้วย '='
    df = df[~df.apply(lambda row: any(isinstance(cell, str) and cell.startswith('=') for cell in row), axis=1)]

    # ตัดแถวแรก 2 แถวออก
    df = df.iloc[1:].reset_index(drop=True)

    # ตั้งชื่อคอลัมน์ใหม่ตามรูปแบบสุดท้ายที่ต้องการ
    df.columns = [
        'TB Account (Old)', 'Account Name', 'C/F Debit', 'C/F Credit',
    ]

    df.to_excel('C:/Users/kc/Desktop/data_csv_to_template/exports/TB(566).xlsx', index=False, engine='openpyxl')
    print("✅ Export Files TB : เรียบร้อย")


def ProcessFile_5B9(df):
    def attach_note_to_previous_row(df):
        """
        ตรวจจับแถวที่มีคำว่า 'หมายเหตุ' ใน column 'D' แล้วนำข้อมูลจากแถวถัดไปที่ 'B' และ 'C' ว่าง
        ไปต่อท้ายใน column 'U' ของแถวก่อนหน้า แล้วลบแถว 'หมายเหตุ' และแถวต่อเนื่องนั้นออก
        """
        df['U'] = None  # เตรียมคอลัมน์ใหม่ไว้เก็บ note

        rows_to_drop = []
        i = 0
        while i < len(df):
            d_val = str(df.iloc[i]['D']) if pd.notna(df.iloc[i]['D']) else ''
            if 'หมายเหตุ' in d_val:
                if i == 0:
                    i += 1
                    continue  # ข้ามถ้าอยู่แถวแรก

                note_lines = []
                j = i + 1
                while j < len(df):
                    b_val = str(df.iloc[j]['B']) if pd.notna(df.iloc[j]['B']) else ''
                    c_val = str(df.iloc[j]['C']) if pd.notna(df.iloc[j]['C']) else ''
                    d_note = str(df.iloc[j]['D']) if pd.notna(df.iloc[j]['D']) else ''

                    if b_val.strip() == '' and c_val.strip() == '' and d_note.strip() != '':
                        note_lines.append(d_note.strip())
                        rows_to_drop.append(j)
                        j += 1
                    else:
                        break

                full_note = "หมายเหตุ:\n" + "\n".join(note_lines)
                df.at[df.index[i - 1], 'U'] = full_note if full_note else None

                rows_to_drop.append(i)
                i = j
            else:
                i += 1

        if rows_to_drop:
            df = df.drop(index=[df.index[i] for i in rows_to_drop]).reset_index(drop=True)

        return df

    df = df.reset_index(drop=True)
    df.columns = list("ABCDEFGHIJKLMNOPQR")[:len(df.columns)]

    # เตรียมคอลัมน์สำหรับ header group
    df['Account1'] = None
    df['Description1'] = None
    df['Depr'] = None

    current_account = None
    current_desc = None
    current_depr = None

    for i, row in df.iterrows():
        val = str(row['B'])
        if val.startswith(('117-', '123-', '124-')):
            # เจอ header ใหม่
            current_account = row['B']
            current_desc = row['C']
            current_depr = row['L']
        else:
            # แถวข้อมูลจริง → เติมค่าล่าสุด
            df.at[i, 'Account1'] = current_account
            df.at[i, 'Description1'] = current_desc
            df.at[i, 'Depr'] = current_depr

    # กรองเฉพาะแถวที่เป็นข้อมูลจริง (มี DocNum)
    df = df[~df['B'].astype(str).str.strip().eq('รวม')]
    df = df[~df['B'].astype(str).str.strip().eq('ทรัพย์สินที่ขายระหว่างงวดบัญชี')]
    df = df[~df['B'].astype(str).str.strip().eq('รวมทรัพย์สินที่ยังเหลืออยู่')]
    df = df[~df['B'].astype(str).str.strip().eq('ทรัพย์สิน')]
    df = df[~df['B'].astype(str).str.startswith('ตัวอย่าง')]
    df = df[~df['B'].astype(str).str.startswith('รายละเอียดค่าเสื่อมราคา')]
    df = df[~df['B'].astype(str).str.startswith('สำหรับงวด')]
    # df = df[~df['G'].astype(str).str.startswith('รวม(B+C)')]
    df = df[~df['D'].astype(str).str.strip().eq('ซื้อ')]
    df = df[~df['F'].astype(str).str.strip().eq('ค่าเสื่อม')]
    df = df.dropna(subset=['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'], how='all')
    df = df[df['Account1'].notna()]

    # 🧠 เรียกใช้ function จัดการหมายเหตุ
    df = attach_note_to_previous_row(df)
    df = df.dropna(subset=['F', 'G', 'H', 'I', 'J'], how='all')
    # จัดคอลัมน์ใหม่ให้อ่านง่าย
    df = df[['Account1', 'Description1', 'Depr', 'B', 'C', 'I', 'N', 'O', 'Q', 'R', 'U']]
    df.columns = [
        'Asset Account', 'Asset Profile Name', 'Depr.Expense Account', 'Asset Code',
        'Asset Name', 'Text_1', 'Asset Start Date', 'Number of Years',
        'Text_2', 'Salvage Value', 'Note'
    ]

    df.to_excel('C:/Users/kc/Desktop/data_csv_to_template/exports/5B9.xlsx', index=False, engine='openpyxl')
    print("✅ Export Files 5B9 : เรียบร้อย")


if __name__ == "__main__":
    print('=============== START PROGRAM =================')
    df_1A2 = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/1A2.CSV', encoding='cp874', header=None, skiprows=6, on_bad_lines='skip')
    ProcessFile_1A2(df_1A2)

    df_116 = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/116.CSV', encoding='cp874', header=None, skiprows=6, on_bad_lines='skip')
    ProcessFile_116(df_116)

    df_3C = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/3C.CSV', encoding='cp874', header=None, skiprows=3, on_bad_lines='skip')
    ProcessFile_3C(df_3C)

    df_2A2 = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/2A2.CSV', encoding='cp874', header=None, skiprows=5, on_bad_lines='skip')
    ProcessFile_2A2(df_2A2)

    df_215 = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/215.csv', encoding='cp874', header=None, skiprows=5, on_bad_lines='skip')
    ProcessFile_215(df_215)

    df_3D = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/3D.csv', encoding='cp874', header=None, skiprows=3, on_bad_lines='skip')
    ProcessFile_3D(df_3D)

    df_TB = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/566.CSV', encoding='cp874', header=None, skiprows=5, on_bad_lines='skip')
    ProcessFile_TB(df_TB)

    df_5B9 = pd.read_csv('C:/Users/kc/Desktop/data_csv_to_template/5B9.CSV', encoding='cp874', header=None, skiprows=3,
                        on_bad_lines='skip')
    ProcessFile_5B9(df_5B9)

    print('================ END PROGRAME ==================')
